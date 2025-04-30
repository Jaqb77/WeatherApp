import os
from fileinput import filename
from openpyxl import Workbook
from openpyxl.reader.excel import load_workbook

def saveToExcel(filename, data):
    if os.path.exists(filename):
        wb = load_workbook(filename)
        ws = wb.active # wybranie aktywnego arkusza
    else:
        wb = Workbook() # stworzenie nowego excela
        ws = wb.active # wybranie aktywnego arkusza
        ws.append(list(data.keys()))

    ws.append(list(data.values()))
    wb.save(filename)






# def create_excel():
#     wb = Workbook()
#     sheet = wb.active
#     sheet.title = 'Temperatury'
#
#     values = [23,21,22,19,86,48,55,33]
#
#     sheet.append(['Pomiar nr.', 'Temperatura', 'Test'])
#
#     for index,value in enumerate(values, start=1):
#         sheet.append([index, value, f'test {index}'])
#     wb.save('testow.xlsx')
#
# create_excel()